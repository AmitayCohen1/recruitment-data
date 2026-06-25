"""
Regenerate every app/data/*.json so the live dashboard is ALIGNED WITH THE
PUBLISHED FILE  public/full-recruitment-data.xlsx  ("נתוני גיוס מלאים").

That file is the source of truth: its combat/officer rates are correctly weighted
by ENLISTEES (שיעור קרבי = Σלוחמים / Σמתגייסים) and it bakes in the right cohort
weights (2024 → תלמידי יב 2021, 2018 → 2015). We therefore read every AGGREGATE
straight from it — no recomputation, nothing calculated — so the site and the
download always show identical numbers.

What comes from where:
  - sectors.json / regions.json : taken directly from the file. The file has only
        two sheets (2018, 2024), so the aggregated views cover exactly those years.
  - recruitment.json / zero-schools.json : raw per-school rows (all years, no
        weighting, no bug) passed through from Recruitment-data-by-school.xlsx.

Inputs (repo root): public/full-recruitment-data.xlsx, Recruitment-data-by-school.xlsx,
                    full_data_with_classification.csv
Run:  python build_app_data.py
"""
import json
import os
import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "app", "data")
PUBLISHED = os.path.join(HERE, "public", "full-recruitment-data.xlsx")  # "נתוני גיוס מלאים"
RECRUIT = os.path.join(HERE, "Recruitment-data-by-school.xlsx")
CLASS_CSV = os.path.join(HERE, "full_data_with_classification.csv")

RAW_YEARS = list(range(2018, 2025))   # raw per-school data spans all years
FILE_YEARS = [2018, 2024]             # the only years the published file covers
SECTORS = ["חילוני", "דתי לאומי", "חרדי", "דרוזי"]  # display names
# the published file labels the Druze group "דרוזי וערבי" (it also covers Arab /
# mixed schools); we match that row but display it as "דרוזי" (see the note in
# the dashboard's disclaimers).
FILE_SECTORS = {"חילוני", "דתי לאומי", "חרדי", "דרוזי וערבי"}
RENAME = {"דרוזי וערבי": "דרוזי"}
REGIONS = ["מרכז", "פריפריה", "כפרי/קיבוצים", "התנחלויות", "ירושלים"]
R_SECTORS = ["הכל", "חילוני", "דתי לאומי", "חרדי"]
GENDERS = [("ז", "בנים", "m"), ("נ", "בנות", "f")]
ALL = "— הכל —"


def r1(x):
    return None if x is None else round(x, 1)


# ---------------------------------------------------------------------------
# Parse the published file
# ---------------------------------------------------------------------------
def read_published():
    """{year: [ {sector, azor, gender, enlist, combat, officer, schools,
    cohort, enlistees, fighters, officers_abs}, ... ]}."""
    wb = openpyxl.load_workbook(PUBLISHED, read_only=True, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        try:
            year = int(sheet)
        except ValueError:
            continue
        rows = []
        for r in wb[sheet].iter_rows(values_only=True):
            if r[0] not in FILE_SECTORS or r[2] not in ("בנים", "בנות"):
                continue
            rows.append({
                "sector": RENAME.get(r[0], r[0]), "azor": r[1], "gender": r[2],
                "enlist": r[3], "combat": r[4], "officer": r[5],
                "schools": int(r[6]) if r[6] is not None else 0,
                "cohort": float(r[7] or 0), "enlistees": float(r[8] or 0),
                "fighters": float(r[9] or 0), "officers_abs": float(r[10] or 0),
            })
        out[year] = rows
    return out


def canonical_region(azor):
    if azor in ("ערי מרכז", "מרכז"):
        return "מרכז"
    if azor in ("ערי פריפריה", "פריפריה"):
        return "פריפריה"
    if azor in ("קיבוצים/כפרי", "פריפריה/כפרי"):
        return "כפרי/קיבוצים"
    if azor == "התנחלויות":
        return "התנחלויות"
    if azor == "ירושלים":
        return "ירושלים"
    return None  # ערי חרדים / ערים מעורבות / אחר → not a geographic region


def group_label(sector, azor):
    if sector == "דרוזי":
        return "דרוזי"
    return f"{sector} - {azor}"


def sector_of(group):
    g = str(group)
    if g.startswith("חרדי"):
        return "חרדי"
    if g.startswith("דתי לאומי"):
        return "דתי לאומי"
    if g.startswith("חילוני"):
        return "חילוני"
    if g in ("דרוזי", "ערבי", "בדואי"):
        return "דרוזי"   # displayed as "דרוזי"; also covers Arab / mixed schools
    return None


# ---------------------------------------------------------------------------
# Raw per-school sources
# ---------------------------------------------------------------------------
def load_classification():
    csv = pd.read_csv(CLASS_CSV)
    csv["school_key"] = csv["school_key"].astype(int)
    cls = csv.drop_duplicates("school_key")[["school_key", "קבוצה"]].copy()
    cls["sector"] = cls["קבוצה"].map(sector_of)
    return cls


def load_recruitment():
    frames = []
    for yr in RAW_YEARS:
        for he, glabel, gcode in GENDERS:
            sheet = f"{yr} {'גברים' if he == 'ז' else 'נשים'}"
            d = pd.read_excel(RECRUIT, sheet_name=sheet, engine="openpyxl").rename(columns={
                "מוסד": "school", "מועצה": "council", "גיוס": "enlist",
                "לחימה": "combat", "קצונה": "officer", "משמעותי": "meaning"})
            d["year"] = yr
            d["gender"] = gcode
            d["school_key"] = d["school_key"].astype(int)
            frames.append(d[["year", "gender", "school_key", "school",
                             "council", "enlist", "combat", "officer", "meaning"]])
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def main():
    pub = read_published()
    cls = load_classification()
    rec = load_recruitment()

    # ---- recruitment.json / zero-schools.json (raw, disjoint) ---------------
    recruitment = [
        {"year": int(r.year), "gender": r.gender, "key": int(r.school_key),
         "school": r.school, "council": r.council,
         "enlist": round(float(r.enlist), 2),
         "combat": None if pd.isna(r.combat) else round(float(r.combat), 2),
         "officer": None if pd.isna(r.officer) else round(float(r.officer), 2),
         "meaning": None if pd.isna(r.meaning) else round(float(r.meaning), 2)}
        for r in rec.itertuples(index=False)
        if not pd.isna(r.enlist) and float(r.enlist) > 0
    ]
    zero = [
        {"y": int(r.year), "g": r.gender, "k": int(r.school_key),
         "s": r.school, "c": r.council}
        for r in rec.itertuples(index=False)
        if not pd.isna(r.enlist) and float(r.enlist) == 0
    ]

    # ---- sectors.json :: byYearSector (straight from the file) --------------
    by_year_sector = []
    for year in FILE_YEARS:
        for row in pub[year]:
            if row["azor"] != ALL:
                continue
            by_year_sector.append({
                "year": year, "sector": row["sector"], "gender": row["gender"],
                "n": row["schools"], "enlist": r1(row["enlist"]),
                "combat": r1(row["combat"]), "officer": r1(row["officer"]),
                "meaning": None,
                # absolute (estimated) counts, straight from the file
                "nCohort": round(row["cohort"]), "nEnlistees": round(row["enlistees"]),
                "nFighters": round(row["fighters"]), "nOfficers": round(row["officers_abs"]),
            })

    # ---- sectors.json :: subgroups (2024, from the file) --------------------
    subgroups = []
    for row in pub[2024]:
        if row["azor"] == ALL:
            continue
        subgroups.append({
            "group": group_label(row["sector"], row["azor"]),
            "sector": row["sector"], "gender": row["gender"], "n": row["schools"],
            "enlist": r1(row["enlist"]), "combat": r1(row["combat"]),
            "officer": r1(row["officer"]), "meaning": None,
        })

    # ---- sectors.json :: schoolSector ---------------------------------------
    school_sector = {
        str(int(r.school_key)): r.sector
        for r in cls.itertuples(index=False)
        if r.sector in SECTORS
    }

    sectors_json = {
        "sectors": SECTORS, "years": FILE_YEARS,
        "byYearSector": by_year_sector, "subgroups": subgroups,
        "schoolSector": school_sector,
        "weighted": True, "weightBase": "2018→2015, 2024→2021",
    }

    # ---- regions.json (from the file) ---------------------------------------
    region_rows = []
    for year in FILE_YEARS:
        for _, glabel, _ in GENDERS:
            grows = [r for r in pub[year] if r["gender"] == glabel and r["azor"] != ALL]
            for r in grows:  # per-sector rows mapped to a geographic region
                region = canonical_region(r["azor"])
                if region is None or r["sector"] not in ("חילוני", "דתי לאומי", "חרדי"):
                    continue
                region_rows.append({
                    "year": year, "region": region, "sector": r["sector"], "gender": glabel,
                    "n": r["schools"], "enlist": r1(r["enlist"]),
                    "combat": r1(r["combat"]), "officer": r1(r["officer"]), "meaning": None,
                })
            for region in REGIONS:  # "הכל" = pool the Jewish sectors via absolutes
                cells = [r for r in grows
                         if canonical_region(r["azor"]) == region
                         and r["sector"] in ("חילוני", "דתי לאומי", "חרדי")]
                if not cells:
                    continue
                cohort = sum(c["cohort"] for c in cells)
                enl = sum(c["enlistees"] for c in cells)
                fig = sum(c["fighters"] for c in cells)
                off = sum(c["officers_abs"] for c in cells)
                region_rows.append({
                    "year": year, "region": region, "sector": "הכל", "gender": glabel,
                    "n": sum(c["schools"] for c in cells),
                    "enlist": r1(100 * enl / cohort) if cohort else None,
                    "combat": r1(100 * fig / enl) if enl else None,
                    "officer": r1(100 * off / enl) if enl else None,
                    "meaning": None,
                })
    regions_json = {
        "regions": REGIONS, "sectors": R_SECTORS, "years": FILE_YEARS, "rows": region_rows,
    }

    # ---- write --------------------------------------------------------------
    def dump(name, obj):
        path = os.path.join(OUT, name)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  wrote {name:22s} ({os.path.getsize(path):,} bytes)")

    print("Writing app/data/:")
    dump("recruitment.json", recruitment)
    dump("zero-schools.json", zero)
    dump("sectors.json", sectors_json)
    dump("regions.json", regions_json)
    print(f"\nrecruitment: {len(recruitment):,} | zero: {len(zero):,} | "
          f"byYearSector: {len(by_year_sector)} | subgroups: {len(subgroups)} | "
          f"regions: {len(region_rows)} | schools: {len(school_sector)}")


if __name__ == "__main__":
    main()
