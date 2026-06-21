"""
Verify the untrusted classification (full_data_with_classification.csv)
against the OFFICIAL Ministry of Education school list (mosdot.xlsx).

Usage:
    1. Download mosdot.xlsx from https://data.gov.il/dataset/mosdot
       and put it in this folder (or pass its path as the first argument).
    2. python3 verify_classification.py [path-to-mosdot.xlsx]

Needs only the standard library + openpyxl (already installed).
Reports:
    - how many of his school classifications match the official פיקוח
    - the 6 hand-overridden schools, official vs his
    - whether the headline sector gaps survive when recomputed
      from the OFFICIAL classification instead of his.
"""
import csv
import sys
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(HERE, "full_data_with_classification.csv")
RECRUIT_PATH = os.path.join(HERE, "Recruitment-data-by-school.xlsx")
ARTICLE_PATH = os.path.join(HERE, "ARTICLE_SUMMARY_TABLES.xlsx")

# schools the script hand-overrides (classify_schools.py lines 97-106)
OVERRIDES = {
    472555: "ממלכתי דתי",
    641365: "ממלכתי דתי",
    441204: "ממלכתי דתי",
    440206: "ממלכתי דתי",
    140954: "חרדי",
    640631: "חרדי",
}


def is_xlsx(path):
    """A real .xlsx is a zip; reject HTML/redirect junk saved with an .xlsx name."""
    try:
        with open(path, "rb") as f:
            return f.read(2) == b"PK"
    except OSError:
        return False


def find_official():
    for cand in sys.argv[1:2] + [
        os.path.join(HERE, "mosdot.xlsx"),
        os.path.join(HERE, "c338378e-93ef-4261-ac66-8d0c487f4b01.xlsx"),
    ]:
        if cand and os.path.exists(cand):
            if is_xlsx(cand):
                return cand
            print(f"  ! {cand} exists but is not a valid .xlsx (probably an HTML/redirect page) — skipping.")
    return None


def col_index(header, *needles):
    """Find a column by exact name, then by substring."""
    for i, h in enumerate(header):
        if h in needles:
            return i
    for i, h in enumerate(header):
        if h and any(n in str(h) for n in needles):
            return i
    return None


def load_official(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_key = col_index(header, "סמל מוסד", "סמל")
    ci_pik = col_index(header, "פיקוח")
    ci_mag = col_index(header, "מגזר")
    if ci_key is None or ci_pik is None:
        print("  columns found:", header)
        raise SystemExit("Could not find 'סמל מוסד' / 'פיקוח' columns in official file.")
    official = {}
    for r in rows:
        try:
            k = int(r[ci_key])
        except (TypeError, ValueError):
            continue
        official.setdefault(
            k,
            {
                "pikuach": (r[ci_pik] or "").strip() if isinstance(r[ci_pik], str) else r[ci_pik],
                "magzar": (r[ci_mag] or "").strip() if ci_mag is not None and isinstance(r[ci_mag], str) else None,
            },
        )
    return official


def find_students():
    for cand in sys.argv[2:3] + [
        os.path.join(HERE, "תלמידי יב.xlsx"),
        os.path.join(HERE, "תלמידים יב.xlsx"),
        os.path.join(HERE, "students.xlsx"),
    ]:
        if cand and os.path.exists(cand) and is_xlsx(cand):
            return cand
    return None


def load_students(path):
    """school_key -> {'boys': n, 'girls': n}, summed across rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_key = col_index(header, "סמל מוסד", "סמל")
    ci_b = col_index(header, "מספר בנים", "בנים")
    ci_g = col_index(header, "מספר בנות", "בנות")
    if ci_key is None or ci_b is None or ci_g is None:
        print("  student columns found:", header)
        return {}
    out = {}
    for r in rows:
        try:
            k = int(r[ci_key])
        except (TypeError, ValueError):
            continue
        b = r[ci_b] if isinstance(r[ci_b], (int, float)) else 0
        g = r[ci_g] if isinstance(r[ci_g], (int, float)) else 0
        d = out.setdefault(k, {"boys": 0, "girls": 0})
        d["boys"] += b or 0
        d["girls"] += g or 0
    return out


def load_his_csv():
    rows = list(csv.DictReader(open(CSV_PATH, encoding="utf-8-sig")))
    his = {}
    recruit = {}
    for r in rows:
        k = int(float(r["school_key"]))
        his.setdefault(k, {"pikuach": (r.get("פיקוח") or "").strip(),
                           "kvutza": (r.get("קבוצה") or "").strip()})
        try:
            recruit[(k, r["מין"])] = float(r["גיוס"])
        except (ValueError, KeyError, TypeError):
            pass
    return his, recruit


def canon_pik(v):
    """Normalize פיקוח to a canonical label. The official file uses abbreviations
    (מ"מ, חמ"ד), his CSV uses full words (ממלכתי, ממלכתי דתי)."""
    if not v:
        return ""
    s = str(v).replace('"', "").replace("'", "").replace("״", "").replace(" ", "").strip()
    if s in ("ממ", "ממלכתי"):
        return "ממלכתי"
    if s in ("חמד", "ממלכתידתי"):
        return "ממלכתי דתי"
    if "חרדי" in s:
        return "חרדי"
    if "מוכר" in s:
        return "מוכר שאינו רשמי"
    if "פטור" in s:
        return "פטור"
    return str(v).strip()


def top_sector(pikuach, magzar):
    if magzar in ("ערבי", "דרוזי", "בדואי"):
        return magzar
    p = canon_pik(pikuach)
    if p == "חרדי":
        return "חרדי"
    if p == "ממלכתי דתי":
        return "דתי לאומי"
    if p == "ממלכתי":
        return "חילוני"
    return "אחר/לא מסווג"


def mean(v):
    v = [x for x in v if x is not None]
    return round(sum(v) / len(v), 1) if v else None


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def sector_stats(year, official, students):
    """For a year, return {(sector,gender): {n, recruit_s/w, combat_s/w}} from raw data,
    classified by the OFFICIAL list. Weighted by gender-specific 12th-grade counts if available."""
    wb = openpyxl.load_workbook(RECRUIT_PATH, read_only=True, data_only=True)
    cells = {}
    for sheet, g, wkey in [("%d גברים" % year, "בנים", "boys"),
                           ("%d נשים" % year, "בנות", "girls")]:
        if sheet not in wb.sheetnames:
            continue
        for i, r in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0 or r[0] is None:
                continue
            k = int(r[0])
            info = official.get(k)
            if not info:
                continue
            sec = top_sector(info["pikuach"], info["magzar"])
            enlist, combat = num(r[4]), num(r[5])
            if enlist is None:
                continue
            w = (students.get(k, {}).get(wkey, 0) or 0) if students else 0
            cells.setdefault((sec, g), []).append((enlist, combat, w))
    out = {}
    for key, vals in cells.items():
        es = [e for e, _, _ in vals]
        cs = [c for _, c, _ in vals if c is not None]
        ws_e = [(e, w) for e, _, w in vals if w > 0]
        ws_c = [(c, w) for _, c, w in vals if c is not None and w > 0]
        def wm(pairs):
            tot = sum(w for _, w in pairs)
            return round(sum(v * w for v, w in pairs) / tot, 1) if tot else None
        out[key] = {
            "n": len(es),
            "recruit_s": mean(es), "recruit_w": wm(ws_e),
            "combat_s": mean(cs), "combat_w": wm(ws_c),
        }
    return out


def read_his_table1():
    """Read his published Table 1 into {(sector,gender): {...}}."""
    if not os.path.exists(ARTICLE_PATH):
        return None
    wb = openpyxl.load_workbook(ARTICLE_PATH, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "Table 1" in s), None)
    if not sheet:
        return None
    rows = list(wb[sheet].iter_rows(values_only=True))
    header = [str(h) for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out = {}
    for r in rows[1:]:
        if r[0] is None:
            continue
        key = (str(r[idx["Sector"]]).strip(), str(r[idx["Gender"]]).strip())
        out[key] = {
            "r2018": num(r[idx.get("2018_Recruitment", -1)]),
            "c2018": num(r[idx.get("2018_Combat", -1)]),
            "r2024": num(r[idx.get("2024_Recruitment", -1)]),
            "c2024": num(r[idx.get("2024_Combat", -1)]),
        }
    return out


def main():
    official_path = find_official()
    if not official_path:
        print("=" * 64)
        print("OFFICIAL FILE NOT FOUND.")
        print("Download mosdot.xlsx from https://data.gov.il/dataset/mosdot")
        print("and place it in this folder, then re-run.")
        print("=" * 64)
        return

    print(f"Official file: {official_path}")
    official = load_official(official_path)
    his, _ = load_his_csv()
    print(f"Official schools: {len(official):,}   |   His schools: {len(his):,}")

    students_path = find_students()
    students = load_students(students_path) if students_path else {}
    if students_path:
        print(f"Student-count file: {students_path}  ({len(students):,} schools) -> WEIGHTED comparison enabled")
    else:
        print("Student-count file: NOT FOUND -> Table-1 check uses SIMPLE means (approx; expect small gaps vs his weighted figures)")

    # --- 1. classification match rate -----------------------------------
    in_both = [k for k in his if k in official]
    match = same = diff = handset = 0
    diffs = []
    for k in in_both:
        op = official[k]["pikuach"]
        hp = his[k]["pikuach"]
        if op in (None, ""):
            handset += 1
            continue
        match += 1
        if canon_pik(op) == canon_pik(hp):
            same += 1
        else:
            diff += 1
            diffs.append((k, canon_pik(op), canon_pik(hp)))

    print("\n" + "=" * 64)
    print("1) CLASSIFICATION MATCH vs OFFICIAL פיקוח")
    print("=" * 64)
    print(f"  His schools also in official list : {len(in_both):,}")
    print(f"  Comparable (official has פיקוח)   : {match:,}")
    print(f"  MATCH                             : {same:,} ({100*same/match:.1f}%)")
    print(f"  DIFFER                            : {diff:,} ({100*diff/match:.1f}%)")
    print(f"  Official blank / hand-set by him  : {handset:,}")
    print(f"  His schools NOT in official list  : {len(his)-len(in_both):,}")

    if diffs:
        print(f"\n  First differences (official -> his):")
        for k, op, hp in diffs[:25]:
            flag = "  <-- OVERRIDE" if k in OVERRIDES else ""
            print(f"    {k}: {op!r} -> {hp!r}{flag}")

    # --- 2. the 6 overrides ---------------------------------------------
    print("\n" + "=" * 64)
    print("2) THE 6 HAND-OVERRIDDEN SCHOOLS")
    print("=" * 64)
    for k, scripted in OVERRIDES.items():
        o = canon_pik(official.get(k, {}).get("pikuach")) or "(not in official)"
        h = canon_pik(his.get(k, {}).get("pikuach")) or "(not in his csv)"
        verdict = "his MATCHES official" if o == h else "his DIFFERS from official"
        print(f"  {k}: official={o!r}  script_wanted={scripted!r}  his_csv={h!r}  -> {verdict}")

    # --- 3. do the headline sector gaps survive? ------------------------
    print("\n" + "=" * 64)
    print("3) HEADLINE SECTORS RECOMPUTED FROM OFFICIAL CLASSIFICATION (2024)")
    print("=" * 64)
    wb = openpyxl.load_workbook(RECRUIT_PATH, read_only=True, data_only=True)
    buckets = {}
    for sheet, g in [("2024 גברים", "בנים"), ("2024 נשים", "בנות")]:
        for i, r in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0 or r[0] is None:
                continue
            k = int(r[0])
            info = official.get(k)
            if not info:
                continue
            sec = top_sector(info["pikuach"], info["magzar"])
            try:
                enlist = float(r[4])
            except (TypeError, ValueError):
                continue
            buckets.setdefault((sec, g), []).append(enlist)
    print(f"  {'sector':14s} {'gender':6s}  n     avg גיוס (official)")
    for sec in ["חילוני", "דתי לאומי", "חרדי", "דרוזי", "ערבי"]:
        for g in ["בנים", "בנות"]:
            v = buckets.get((sec, g), [])
            if v:
                print(f"  {sec:14s} {g:6s}  {len(v):<4d}  {mean(v)}%")
    print("\n  Compare these to his ARTICLE_SUMMARY_TABLES Table 1.")
    print("  If the gaps look the same -> his story holds.")

    # --- 4. direct diff of his Table 1 numbers --------------------------
    his_t1 = read_his_table1()
    if his_t1 is None:
        print("\n(ARTICLE_SUMMARY_TABLES.xlsx not found — skipping Table-1 diff.)")
        return

    weighted = bool(students)
    tol = 1.5 if weighted else 6.0  # weighted should match tightly; simple-mean is approximate
    mode = "WEIGHTED" if weighted else "SIMPLE (approx)"
    pick = "recruit_w" if weighted else "recruit_s"
    pickc = "combat_w" if weighted else "combat_s"

    print("\n" + "=" * 64)
    print(f"4) DIRECT DIFF OF HIS TABLE 1  (mode: {mode}, flag if |Δ| > {tol})")
    print("=" * 64)
    mine = {2018: sector_stats(2018, official, students),
            2024: sector_stats(2024, official, students)}

    print(f"  {'sector':12s}{'g':5s}{'metric':10s}{'his':>8s}{'mine':>8s}{'Δ':>7s}  flag")
    flags = 0
    for (sec, g), hv in sorted(his_t1.items()):
        for yr, hk, mk, lbl in [
            (2024, "r2024", pick, "גיוס24"),
            (2024, "c2024", pickc, "לחימה24"),
            (2018, "r2018", pick, "גיוס18"),
            (2018, "c2018", pickc, "לחימה18"),
        ]:
            h = hv.get(hk)
            m = mine[yr].get((sec, g), {}).get(mk)
            if h is None or m is None:
                continue
            d = round(h - m, 1)
            bad = abs(d) > tol
            flags += bad
            print(f"  {sec:12s}{g:5s}{lbl:10s}{h:8.1f}{m:8.1f}{d:7.1f}  {'<-- FLAG' if bad else 'ok'}")

    print(f"\n  {flags} cell(s) flagged.")
    if weighted:
        print("  Weighted mode: flags mean his published numbers don't reproduce -> investigate.")
    else:
        print("  Simple mode: small gaps are expected (his are weighted). Large gaps still matter.")
        print("  Add the 'תלמידי יב' student file for an exact weighted check.")


if __name__ == "__main__":
    main()
